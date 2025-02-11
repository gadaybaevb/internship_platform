from django.shortcuts import render, get_object_or_404, redirect
from openpyxl.cell import MergedCell
from datetime import date
from .models import Material, Position, Internship, StageProgress, MaterialProgress
from departments.models import Department
from django.contrib.auth import get_user_model
from .forms import MaterialForm, ReviewForm, AddInternForm, MentorReviewForm
from django.db.models import Q, Count
from django.core.paginator import Paginator
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib import messages
from users.models import CustomUser
from django.core.files.storage import default_storage
from .utils import create_stage_progress
from datetime import timedelta, datetime
from django.utils import timezone
from notifications.models import Notification
from tests.models import Test, TestResult
from django.utils.dateparse import parse_date
from django.db.models import Avg, F
from django.contrib.messages import get_messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
import pandas as pd
from unidecode import unidecode
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth.decorators import user_passes_test

User = get_user_model()


@login_required
def material_list(request):
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    position_filter = request.GET.get('position', '')
    stage_filter = request.GET.get('stage', '')

    # Получаем все материалы
    materials = Material.objects.all()

    # Фильтрация по поисковому запросу
    if search_query:
        materials = materials.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(position__name__icontains=search_query)
        )

    # Фильтрация по департаменту
    if department_filter:
        materials = materials.filter(position__department__id=department_filter)

    # Фильтрация по позиции
    if position_filter:
        materials = materials.filter(position__id=position_filter)

    # Фильтрация по этапу
    if stage_filter:
        materials = materials.filter(stage=stage_filter)

    # Пагинация по 50 записей
    paginator = Paginator(materials, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Получаем списки департаментов и позиций для фильтрации
    departments = Department.objects.all()
    positions = Position.objects.all()

    # Подсчёт общего количества материалов
    total_materials = materials.count()

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'departments': departments,
        'positions': positions,
        'department_filter': department_filter,
        'position_filter': position_filter,
        'stage_filter': stage_filter,
        'total_materials': total_materials,  # Количество материалов
    }
    return render(request, 'material_list.html', context)


@login_required
def material_create(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES)
        if form.is_valid():
            # Проверка на наличие дубля
            title = form.cleaned_data['title']
            position = form.cleaned_data['position']
            stage = form.cleaned_data['stage']

            if Material.objects.filter(title=title, position=position, stage=stage).exists():
                # Если такой материал уже существует, выводим сообщение об ошибке
                messages.error(request, 'Материал с таким названием, позицией и этапом уже существует.')
            else:
                # Если дубля нет, сохраняем материал
                form.save()
                messages.success(request, 'Материал успешно добавлен.')
                return redirect('material_list')
    else:
        form = MaterialForm()

    # Получаем только сообщения, которые относятся к этому представлению
    storage = get_messages(request)
    relevant_messages = []
    for message in storage:
        if message.level == messages.ERROR or message.level == messages.SUCCESS:
            relevant_messages.append(message)

    return render(request, 'material_form.html', {'form': form})


@login_required
def material_edit(request, material_id):
    material = get_object_or_404(Material, id=material_id)
    old_file = material.file  # Store the original file path before saving

    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES, instance=material)
        if form.is_valid():
            # Check if the file field is being cleared
            if 'file-clear' in request.POST and old_file:
                if default_storage.exists(old_file.name):
                    default_storage.delete(old_file.name)  # Delete old file if it exists
            form.save()
            return redirect('material_list')
    else:
        form = MaterialForm(instance=material)

    return render(request, 'material_form.html', {'form': form, 'material': material})


@login_required
def material_delete(request, material_id):
    material = get_object_or_404(Material, id=material_id)

    if request.method == 'POST':
        material.delete()
        return redirect('material_list')

    return render(request, 'material_confirm_delete.html', {'material': material})


@staff_member_required
def assign_mentor(request, internship_id):
    internship = get_object_or_404(Internship, id=internship_id)

    if request.method == 'POST':
        mentor_id = request.POST.get('mentor')
        position_id = request.POST.get('position')

        mentor = CustomUser.objects.get(id=mentor_id)
        position = Position.objects.get(id=position_id)

        internship.mentor = mentor
        internship.position = position  # Присваиваем позицию стажировке
        internship.save()

        # Обновляем позицию стажёра в модели CustomUser
        internship.intern.position = position
        internship.intern.save()

        # Создаем этапы стажировки для стажёра, если их ещё нет
        if not StageProgress.objects.filter(intern=internship.intern, position=position).exists():
            create_stage_progress(internship)  # Создаём этапы для стажировки

        messages.success(request, f"Ментор {mentor.username} назначен для стажера {internship.intern.username}, позиция {position.name}.")

        message = f"Ментор {mentor.username} назначен для стажера {internship.intern.username}, позиция {position.name}."
        Notification.objects.create(user=internship.intern, message=message)

        return redirect('internship_list')

    mentors = CustomUser.objects.filter(role='mentor')
    positions = Position.objects.all()  # Получаем все доступные позиции

    context = {
        'internship': internship,
        'mentors': mentors,
        'positions': positions,  # Передаем позиции в шаблон
    }
    return render(request, 'assign_mentor.html', context)


@staff_member_required
def add_intern(request):
    if request.method == 'POST':
        form = AddInternForm(request.POST)
        if form.is_valid():
            # Получаем объекты сразу из формы
            intern = form.cleaned_data['intern']
            mentor = form.cleaned_data['mentor']
            position = form.cleaned_data['position']

            # Создаем новую стажировку
            internship = Internship.objects.create(
                intern=intern,
                mentor=mentor,
                position=position
            )

            messages.success(request, f"Стажировка для {intern.full_name} успешно создана с ментором {mentor.full_name}.")
            return redirect('internship_list')
    else:
        form = AddInternForm()

    return render(request, 'add_intern.html', {'form': form})


@login_required
def delete_internship(request, internship_id):
    internship = get_object_or_404(Internship, id=internship_id)
    internship.delete()
    messages.success(request, "Стажировка успешно удалена.")
    return redirect('internship_list')


@login_required
def internship_list(request):
    search_query = request.GET.get('search', '')  # Получаем запрос поиска
    internships = Internship.objects.all()

    # Если есть запрос на поиск, фильтруем по имени стажера
    if search_query:
        internships = internships.filter(Q(intern__username__icontains=search_query))

    # Пагинация на 50 записей
    paginator = Paginator(internships, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'internship_list.html', {'page_obj': page_obj, 'search_query': search_query})


@login_required
def dashboard(request):
    user = request.user
    check_deadlines(user)  # Проверяем дедлайны

    form = None  # Инициализация переменной `form`

    if user.role == 'admin':
        internships = Internship.objects.select_related('intern', 'mentor', 'position').all()
    elif user.role == 'mentor':
        internships = Internship.objects.select_related('intern', 'position').filter(mentor=user)

        for internship in internships:
            pending_materials_count = MaterialProgress.objects.filter(
                intern=internship.intern,
                status='pending'
            ).count()
            internship.pending_materials_count = pending_materials_count
            internship.mentor_review_exists = bool(internship.mentor_feedback)
            internship.intern_feedback_exists = bool(internship.intern_feedback)
            internship.is_internship_completed = internship.is_completed()

        # Обработка формы отзыва только для отзыва ментора
        if request.method == 'POST' and 'internship_id' in request.POST:
            internship_id = request.POST.get('internship_id')
            internship = get_object_or_404(Internship, id=internship_id)
            form = MentorReviewForm(request.POST, instance=internship)  # Используем форму только для отзыва ментора
            if form.is_valid():
                form.save()
                messages.success(request, 'Ваш отзыв был добавлен.')
                return redirect('dashboard')

        form = MentorReviewForm()  # Пустая форма только для отзыва ментора

    else:
        stage_progress = StageProgress.objects.filter(intern=user).order_by('stage')
        internship = Internship.objects.filter(intern=user).first()
        if not internship:
            messages.error(request, "У вас нет активной стажировки. Пожалуйста, свяжитесь с администратором.")
            return render(request, 'intern_dashboard.html', {'stage_progress': stage_progress})
        return render(request, 'intern_dashboard.html', {'stage_progress': stage_progress})

    paginator = Paginator(internships, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'mentor_admin_dashboard.html', {
        'page_obj': page_obj,
        'role': user.role,
        'form': form,
    })


@login_required
def update_stage_progress(request, stage_id):
    stage = get_object_or_404(StageProgress, id=stage_id)

    # Только ментор может обновить этап
    if request.user.role == 'mentor':
        stage.completed = True
        stage.completion_date = timezone.now()  # Добавляем дату завершения
        stage.save()
        # Сообщение для уведомления
        messages.success(request, f"Этап {stage.stage} для {stage.intern.full_name} отмечен как завершённый.")

        # Уведомляем администратора
        admin_message = f"Стажер {stage.intern.full_name} завершил этап {stage.stage}."
        for admin in CustomUser.objects.filter(role='admin'):
            Notification.objects.create(user=admin, message=admin_message)

    return redirect('dashboard')


@login_required
def intern_materials(request):
    intern = request.user
    internship = Internship.objects.filter(intern=intern).first()

    # Проверяем, есть ли уже отзыв от стажера
    intern_feedback = internship.intern_feedback if internship else None

    # Создаем форму для отзыва, но показываем ее только при отсутствии отзыва
    if request.method == 'POST' and not intern_feedback:
        form = ReviewForm(request.POST, instance=internship)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ваш отзыв был отправлен.')
            return redirect('intern_materials')
    else:
        form = ReviewForm(instance=internship)

    # Если стажировка не найдена, возвращаем пустые данные
    if not internship:
        return render(request, 'intern_materials.html', {'materials': [], 'progress_summary': {}})

    # Дата завершения стажировки
    start_date = internship.start_date
    end_date = start_date + timedelta(days=intern.position.duration_days)
    end_date_time_aware = timezone.make_aware(datetime.combine(end_date, datetime.min.time()))
    time_left = max((end_date_time_aware - timezone.now()).days, 0)

    # Разделяем материалы на завершенные и незавершенные
    completed_materials = []
    pending_materials = []
    materials = Material.objects.filter(position=intern.position).order_by('stage')

    for material in materials:
        material_progress = MaterialProgress.objects.filter(intern=intern, material=material).first()
        # Проверяем статус и добавляем материал в нужный список
        if material_progress:
            if material_progress.completed:
                completed_materials.append({'material': material, 'status': 'completed'})
            elif material_progress.status == 'pending':
                pending_materials.append({'material': material, 'status': 'pending', 'progress_id': material_progress.id})
            else:
                pending_materials.append({'material': material, 'status': 'not_started'})
        else:
            # Если прогресса еще нет, материал не завершен и статус по умолчанию "not_started"
            pending_materials.append({'material': material, 'status': 'not_started'})

    # Подсчет завершенных и оставшихся материалов для каждого этапа
    stage_1_materials = Material.objects.filter(position=intern.position, stage=1)
    stage_1_materials_count = stage_1_materials.count()
    completed_stage_1_count = MaterialProgress.objects.filter(intern=intern, material__in=stage_1_materials, completed=True).count()
    stage_1_remaining = stage_1_materials_count - completed_stage_1_count

    stage_2_materials = Material.objects.filter(position=intern.position, stage=2)
    stage_2_materials_count = stage_2_materials.count()
    completed_stage_2_count = MaterialProgress.objects.filter(intern=intern, material__in=stage_2_materials, completed=True).count()
    stage_2_remaining = stage_2_materials_count - completed_stage_2_count

    # Проверка и обновление завершения этапов
    stage_1_completed = completed_stage_1_count == stage_1_materials_count
    stage_2_completed = completed_stage_2_count == stage_2_materials_count

    # Обновляем статус этапов в StageProgress для каждого этапа, если этап завершен
    if stage_1_completed:
        stage_progress_1, created = StageProgress.objects.get_or_create(intern=intern, stage=1, position=intern.position)
        if not stage_progress_1.completed:
            stage_progress_1.completed = True
            stage_progress_1.completion_date = timezone.now()
            stage_progress_1.save()

    if stage_2_completed:
        stage_progress_2, created = StageProgress.objects.get_or_create(intern=intern, stage=2, position=intern.position)
        if not stage_progress_2.completed:
            stage_progress_2.completed = True
            stage_progress_2.completion_date = timezone.now()
            stage_progress_2.save()

    # Получаем промежуточный и финальный тесты и их результаты
    midterm_test = Test.objects.filter(position=intern.position, stage_number=1).first()
    final_test = Test.objects.filter(position=intern.position, stage_number=2).first()
    midterm_test_result = TestResult.objects.filter(user=intern, test=midterm_test).first() if midterm_test else None
    final_test_result = TestResult.objects.filter(user=intern, test=final_test).first() if final_test else None

    # Условие для отображения кнопок тестов
    show_midterm_test_button = (not midterm_test_result) and midterm_test and stage_1_completed
    show_final_test_button = (not final_test_result) and final_test and stage_2_completed

    # Условие для отображения формы отзыва стажера
    show_feedback_form = internship.is_completed() and not intern_feedback

    # Передаем все данные в шаблон
    return render(request, 'intern_materials.html', {
        'pending_materials': pending_materials,
        'completed_materials': completed_materials,
        'progress_summary': {
            'total': len(materials),
            'completed': len(completed_materials),
            'remaining': len(materials) - len(completed_materials),
            'time_left': time_left,
        },
        'form': form,
        'intern': intern,
        'internship': internship,
        'stage_1_completed': stage_1_completed,
        'stage_2_completed': stage_2_completed,
        'show_midterm_test_button': show_midterm_test_button,
        'show_final_test_button': show_final_test_button,
        'midterm_test': midterm_test,
        'final_test': final_test,
        'midterm_test_result': midterm_test_result,
        'final_test_result': final_test_result,
        'show_feedback_form': show_feedback_form,
        'intern_feedback': intern_feedback,
        'stage_1_remaining': stage_1_remaining,
        'stage_2_remaining': stage_2_remaining,
    })


def check_deadlines(user):
    # Получаем все активные этапы стажировки для стажера
    stages_in_progress = StageProgress.objects.filter(intern=user, completed=False)

    for stage in stages_in_progress:
        # Проверяем наличие активной стажировки у стажера
        internship = Internship.objects.filter(intern=stage.intern).first()

        if internship:
            end_date = internship.start_date + timedelta(days=stage.position.duration_days)
            days_left = (end_date - timezone.now().date()).days

            # Если осталось 7 дней до конца этапа или стажировки
            if days_left <= 7 and not stage.intern.notifications.filter(message__contains="7 дней").exists():
                message = f"До завершения этапа {stage.stage} осталось {days_left} дней."
                Notification.objects.create(user=stage.intern, message=message)

            # Если этап уже просрочен
            if days_left <= 0 and not stage.intern.notifications.filter(message__contains="просрочен").exists():
                message = f"Этап {stage.stage} просрочен."
                Notification.objects.create(user=stage.intern, message=message)
        else:
            # Если стажировка не найдена, можно выводить сообщение
            message = "Стажировка не найдена. Пожалуйста, свяжитесь с администратором."
            Notification.objects.create(user=stage.intern, message=message)


@login_required
def mark_material_completed(request, material_id):
    material = get_object_or_404(Material, id=material_id)
    intern = request.user

    # Получаем или создаем прогресс для материала
    material_progress, created = MaterialProgress.objects.get_or_create(
        intern=intern,
        material=material,
        defaults={'status': 'not_started'}
    )

    if request.method == 'POST':
        feedback = request.POST.get('feedback', '').strip()  # Получаем отзыв из формы
        if feedback:
            # Сохраняем статус и отзыв
            material_progress.feedback = feedback
            material_progress.status = 'pending'
            material_progress.completed = False  # Ожидаем подтверждения от ментора
            material_progress.completion_date = timezone.now()  # Устанавливаем дату завершения
            material_progress.save()

            # Отправляем уведомление ментору
            internship = Internship.objects.filter(intern=intern).first()
            mentor = internship.mentor if internship else None
            if mentor:
                Notification.objects.create(
                    user=mentor,
                    message=f"Стажер {intern.full_name} отметил материал '{material.title}' как завершённый. Отзыв: {feedback}. Требуется подтверждение."
                )

            messages.success(request, 'Материал был отмечен как завершённый, ожидается подтверждение ментора.')
        else:
            messages.error(request, 'Пожалуйста, оставьте отзыв о материале.')

    return redirect('intern_materials')


@login_required
def mentor_view_intern_materials(request, intern_id):
    intern = get_object_or_404(CustomUser, id=intern_id)

    materials = Material.objects.filter(position=intern.position).order_by('stage')
    material_list = []

    for material in materials:
        material_progress = MaterialProgress.objects.filter(intern=intern, material=material).first()

        if material_progress:
            if material_progress.completed:
                status = 'Завершён'
            else:
                status = 'Ожидание'
        else:
            status = 'Не завершён'

        material_list.append({
            'material': material,
            'status': status,
            'progress_id': material_progress.id if material_progress else None,
            'feedback': material_progress.feedback if material_progress else None,  # Отзыв интерна
        })

    return render(request, 'mentor_view_intern_materials.html', {
        'intern': intern,
        'materials': material_list
    })


@login_required
def confirm_material_completion(request, progress_id):
    material_progress = get_object_or_404(MaterialProgress, id=progress_id)

    if request.user.role == 'mentor' and request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            material_progress.status = 'completed'
            material_progress.completed = True
            material_progress.confirmation_date = timezone.now()
            material_progress.save()
            print("Material approved")  # Отладка
            messages.success(request, 'Материал успешно подтверждён как завершённый.')
        elif action == 'reject':
            rejection_reason = request.POST.get('rejection_reason', '').strip()
            print(f"Rejection reason: {rejection_reason}")  # Отладка
            if rejection_reason:
                Notification.objects.create(
                    user=material_progress.intern,
                    message=f"Ваш материал '{material_progress.material.title}' отклонён. Причина: {rejection_reason}"
                )
                material_progress.status = 'not_started'
                material_progress.feedback = None
                material_progress.save()
                print("Material rejected")  # Отладка
                messages.success(request, 'Материал отклонён, уведомление отправлено стажёру.')
            else:
                print("Rejection reason not provided")  # Отладка
                messages.error(request, 'Пожалуйста, укажите причину отклонения материала.')
    else:
        print("Invalid user role or request method")  # Отладка

    return redirect('dashboard')


@login_required
def intern_report(request, intern_id):
    intern = get_object_or_404(CustomUser, id=intern_id)
    internship = get_object_or_404(Internship, intern=intern)

    # Получаем все тесты, которые прошел стажер
    test_results = TestResult.objects.filter(user=intern)

    # Дата добавления в систему (это поле должно быть в CustomUser)
    date_added = intern.date_joined

    # Дата завершения стажировки (дата последнего теста)
    if test_results.exists():
        completion_date = test_results.order_by('-completed_at').first().completed_at
    else:
        completion_date = None

    # Ментор стажера
    mentor = internship.mentor

    # Получаем отзывы (если они были оставлены)
    intern_feedback = internship.intern_feedback
    mentor_feedback = internship.mentor_feedback

    return render(request, 'intern_report.html', {
        'intern': intern,
        'internship': internship,
        'test_results': test_results,
        'date_added': date_added,  # Дата создания пользователя
        'completion_date': completion_date,
        'mentor': mentor,
        'intern_feedback': intern_feedback,  # Добавляем отзыв стажера
        'mentor_feedback': mentor_feedback,  # Добавляем отзыв ментора
    })


@staff_member_required
def reports_view(request):
    return render(request, 'reports.html')


@staff_member_required
def test_reports_view(request):
    # Получаем параметры фильтрации
    search_query = request.GET.get('search', '')
    test_query = request.GET.get('test', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Основной запрос для фильтрации
    test_results = TestResult.objects.all().order_by('-completed_at')

    # Фильтрация по ФИО стажера или по названию теста
    if search_query:
        test_results = test_results.filter(Q(user__full_name__icontains=search_query) | Q(user__username__icontains=search_query))

    if test_query:
        test_results = test_results.filter(test__title__icontains=test_query)

    # Фильтрация по дате
    if start_date:
        test_results = test_results.filter(completed_at__gte=parse_date(start_date))
    if end_date:
        test_results = test_results.filter(completed_at__lte=parse_date(end_date))

    # Пагинация результатов
    paginator = Paginator(test_results, 10)  # 10 результатов на страницу
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'test_reports.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'test_query': test_query,
        'start_date': start_date,
        'end_date': end_date,
    })


@staff_member_required
def completed_internships_report(request):
    # Получаем параметры фильтрации
    search_query = request.GET.get('search', '')
    position_query = request.GET.get('position', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    # Основной запрос для выборки завершенных стажировок
    internships = Internship.objects.all()

    # Фильтрация по ФИО стажера
    if search_query:
        internships = internships.filter(
            Q(intern__full_name__icontains=search_query) |
            Q(intern__username__icontains=search_query)
        )

    # Фильтрация по позиции
    if position_query:
        internships = internships.filter(position__name__icontains=position_query)

    # Фильтрация по дате начала и завершения стажировки
    if start_date:
        internships = internships.filter(start_date__gte=parse_date(start_date))
    if end_date:
        internships = internships.filter(start_date__lte=parse_date(end_date))

    # Фильтруем только завершенные стажировки
    completed_internships = []
    for internship in internships:
        completed_stages = StageProgress.objects.filter(intern=internship.intern, completed=True)
        all_tests = TestResult.objects.filter(user=internship.intern)
        if completed_stages.exists() and all_tests.exists():
            internship.completed_at = completed_stages.order_by('-completion_date').first().completion_date
            internship.test_results = all_tests  # Добавляем результаты тестов
            completed_internships.append(internship)

    # Пагинация на 10 записей на страницу
    paginator = Paginator(completed_internships, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'completed_internships_report.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'position_query': position_query,
        'start_date': start_date,
        'end_date': end_date,
    })


@staff_member_required
def mentor_report(request):
    department_query = request.GET.get('department', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    mentors = CustomUser.objects.filter(role='mentor')

    if department_query:
        mentors = mentors.filter(department__name__icontains=department_query)

    if start_date and end_date:
        mentors = mentors.filter(intern_internships__start_date__range=[start_date, end_date])

    mentor_stats = []

    for mentor in mentors:
        # Количество стажеров под руководством ментора
        internships = Internship.objects.filter(mentor=mentor)
        total_interns = internships.count()

        # Количество завершенных стажировок
        completed_internships = internships.filter(intern__stage_progresses__completed=True).distinct().count()

        # Количество стажировок в процессе
        in_progress_internships = total_interns - completed_internships

        # Среднее время подтверждения материалов
        avg_confirmation_time = MaterialProgress.objects.filter(
            intern__in=internships.values_list('intern', flat=True),
            status='completed'
        ).aggregate(avg_time=Avg(F('confirmation_date') - F('completion_date')))

        # Процент успешности по тестам (используем related_name 'test_results' в модели TestResult)
        total_tests = TestResult.objects.filter(user__in=internships.values_list('intern', flat=True)).count()
        successful_tests = TestResult.objects.filter(user__in=internships.values_list('intern', flat=True), score__gte=60).count()  # Считаем сданные тесты

        if total_tests > 0:
            test_success_rate = (successful_tests / total_tests) * 100
        else:
            test_success_rate = 0

        mentor_stats.append({
            'mentor': mentor,
            'total_interns': total_interns,
            'completed_internships': completed_internships,
            'in_progress_internships': in_progress_internships,
            'test_success_rate': test_success_rate,
            'avg_confirmation_time': avg_confirmation_time['avg_time'] or timedelta(0),
        })

    return render(request, 'mentor_report.html', {
        'mentor_stats': mentor_stats,
        'department_query': department_query,
        'start_date': start_date,
        'end_date': end_date,
    })


@staff_member_required
def department_materials_report(request):
    # Фильтрация по департаменту
    department_query = request.GET.get('department', '')

    departments = Department.objects.all()

    if department_query:
        departments = departments.filter(name__icontains=department_query)

    department_stats = []

    for department in departments:
        # Позиции в этом департаменте
        positions = Position.objects.filter(department=department)

        position_data = []

        for position in positions:
            # Количество материалов для каждой позиции
            materials_count = Material.objects.filter(position=position).count()

            # Количество тестов для каждой позиции
            tests_count = Test.objects.filter(position=position).count()

            # Количество материалов для каждого этапа
            stages = Material.objects.filter(position=position).values('stage').annotate(
                materials_per_stage=Count('id'))

            # Формируем данные для каждой позиции
            position_data.append({
                'position': position,
                'materials_count': materials_count,
                'tests_count': tests_count,
                'stages': stages,
            })

        department_stats.append({
            'department': department,
            'positions': position_data,
        })

    return render(request, 'department_materials_report.html', {
        'department_stats': department_stats,
        'department_query': department_query,
    })


@staff_member_required
def weekly_report(request):
    # Выбираем всех стажеров
    interns = Internship.objects.all()

    # Создаем список для данных отчета
    report_data = []

    for index, internship in enumerate(interns, start=1):
        intern = internship.intern
        position = internship.position
        department = position.department.name if position and position.department else "No Department"
        supervisor = internship.mentor

        # Подсчет материалов
        total_materials = Material.objects.filter(position=position).count()  # Общее количество материалов
        completed_materials = MaterialProgress.objects.filter(intern=intern, completed=True).count()  # Количество завершенных материалов

        # Отладочная информация

        # Проверяем, что есть данные
        all_test_results = TestResult.objects.filter(user=intern)


        # Попробуем найти результаты по одному полю, чтобы отладить
        if all_test_results.filter(test__position=position).exists():
            print("Data exists for intern and position.")

        mid_test_result = all_test_results.filter(test__position=position, test__stage_number=1).first()
        final_test_result = all_test_results.filter(test__position=position, test__stage_number=2).first()



        # Добавление данных в отчет
        report_data.append({
            '№': index,  # Нумерация
            'Employee': intern.full_name,
            'Department': department,  # Новый столбец "Department"
            'Position': position.name if position else 'Нет позиции',
            'Hiring date': internship.start_date.strftime('%d.%m.%Y'),
            'Probation ending day': (internship.start_date + timezone.timedelta(days=90)).strftime('%d.%m.%Y'),
            'Supervisor': supervisor.full_name if supervisor else "No Supervisor",
            'Status': "on a probation" if completed_materials < total_materials else "completed",
            'Materials passed': f"{completed_materials}/{total_materials}",
            'Mid test': mid_test_result.score if mid_test_result else "Not taken",
            'Final test': final_test_result.score if final_test_result else "Not taken"
        })

    # Создаем DataFrame
    df = pd.DataFrame(report_data)

    # Устанавливаем заголовок
    current_date = timezone.now().date().strftime('%d.%m.%Y')
    header = f"Недельный отчет на {current_date}"

    # Настраиваем ответ для скачивания файла
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="weekly_report_{current_date}.xlsx"'

    # Сохранение в Excel с учетом заголовка и двух пустых строк
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        # Создаем новый лист
        sheet_name = 'Weekly Report'
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)  # Записываем данные с 4-й строки

        # Получаем книгу и лист для добавления заголовка и пустых строк
        workbook = writer.book
        sheet = workbook[sheet_name]

        # Добавляем заголовок и объединяем его ячейки
        sheet['A1'] = header
        sheet.merge_cells('A1:L1')  # Объединяем ячейки для заголовка по ширине таблицы
        sheet['A1'].font = Font(size=14, bold=True)  # Жирный шрифт для заголовка

        # Применяем жирный шрифт ко всем заголовкам столбцов
        for cell in sheet[4]:  # Заголовки находятся в 4-й строке (индекс строки смещен на две пустые строки и заголовок)
            cell.font = Font(bold=True)

        # Устанавливаем автоширину для всех столбцов, кроме первого
        for i, column_cells in enumerate(sheet.columns):
            if i == 0:  # Пропускаем первый столбец (№)
                sheet.column_dimensions[column_cells[0].column_letter].width = 5  # Устанавливаем фиксированную ширину для первого столбца
                continue

            # Пропускаем объединенные ячейки и находим первую не объединенную ячейку для получения буквы столбца
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break

            # Для остальных столбцов вычисляем автоширину
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = max_length + 2  # Добавляем немного отступа
            sheet.column_dimensions[column_letter].width = adjusted_width

    return response


@login_required
def intern_report_export(request, intern_id):
    # Получаем стажера
    intern = get_object_or_404(CustomUser, id=intern_id)

    # Получаем завершенные материалы для стажера
    progress_records = MaterialProgress.objects.filter(intern=intern, completed=True)
    if not progress_records:
        return HttpResponse('<script>alert("Стажер не завершил ни одного материала."); window.history.back();</script>')

    # Подготавливаем данные для отчета с нумерацией
    report_data = []
    for index, record in enumerate(progress_records, start=1):
        completion_time = record.completion_date
        confirmation_time = record.confirmation_date
        time_difference = (confirmation_time - completion_time) if confirmation_time and completion_time else None

        report_data.append({
            '№': index,  # Нумерация
            'Стажер': intern.full_name,
            'Материал': record.material.title if record.material else "Не указан",
            'Отзыв': record.feedback if record.feedback else "Отзыв отсутствует",
            'Дата завершения': completion_time.strftime('%Y-%m-%d %H:%M') if completion_time else "Не завершен",
            'Дата подтверждения': confirmation_time.strftime('%Y-%m-%d %H:%M') if confirmation_time else "Не подтвержден",
            'Время на подтверждение': str(time_difference) if time_difference else "Нет данных"
        })

    # Создаем DataFrame для отчета
    df = pd.DataFrame(report_data)

    # Преобразуем имя стажера в латиницу
    sanitized_name = unidecode(intern.full_name).strip().replace(" ", "_").replace("/", "_")
    today = date.today().strftime('%Y-%m-%d')
    filename = f"{sanitized_name}_{today}.xlsx"

    # Создаем объект ответа
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename=\"{filename}\"; filename*=UTF-8''{filename}"

    # Отладка
    print(f"Generated filename: {filename}")

    # Сохраняем данные в Excel и настраиваем стили
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Отчет по стажеру', startrow=3)  # Данные начинаются с 4-й строки

        # Получаем книгу и лист для добавления стилей
        workbook = writer.book
        sheet = workbook['Отчет по стажеру']

        # Настройки стилей
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")
        cell_font = Font(size=12)
        cell_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Добавляем заголовок и объединяем ячейки
        title = f"Отчет стажировки ({intern.full_name}) на {today}"
        sheet['A1'] = title
        sheet.merge_cells('A1:H1')  # Объединяем ячейки по ширине таблицы
        sheet['A1'].font = header_font
        sheet['A1'].alignment = alignment

        # Применяем стили к заголовкам таблицы (4-я строка)
        for cell in sheet[4]:  # Заголовки столбцов
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        # Применяем стили к ячейкам данных
        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = cell_font
                cell.fill = cell_fill
                cell.alignment = alignment
                cell.border = thin_border

                # Специальное форматирование для столбца "Отзывы" (столбец D)
                if cell.column == 4:  # Номер столбца с "Отзывы"
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if cell.column == 3:  # Для столбца "Материал"
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Устанавливаем автоширину для всех столбцов, кроме столбца "Отзывы"
        for col_num, column_cells in enumerate(sheet.columns, 1):
            if col_num == 4:  # Для столбца "Отзывы"
                sheet.column_dimensions[get_column_letter(col_num)].width = 60
            elif col_num == 3:  # Для столбца "Материал"
                sheet.column_dimensions[get_column_letter(col_num)].width = 60
            else:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                adjusted_width = max_length + 2
                sheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

    return response


@login_required
def active_interns_list(request):
    # Получаем все стажировки
    internships = Internship.objects.all().select_related('intern', 'position', 'mentor')

    # Добавляем данные о количестве материалов для каждого стажера
    internships_data = []
    for internship in internships:
        # Общее количество материалов для позиции стажера
        total_materials = Material.objects.filter(position=internship.position).count()

        # Количество завершенных материалов для стажера
        completed_materials = MaterialProgress.objects.filter(intern=internship.intern, completed=True).count()

        # Добавляем стажировку и её данные в список
        internships_data.append({
            'internship': internship,
            'total_materials': total_materials,
            'completed_materials': completed_materials,
        })

    return render(request, 'active_interns_list.html', {
        'internships_data': internships_data,
    })